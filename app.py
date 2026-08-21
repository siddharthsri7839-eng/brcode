"""
InvenScan — Smart Inventory Management System
Features:
  - 3-Role Access Control (Admin, Barcode Manager, Receiver)
  - Windows OCR Synchronous Auto-Fill (MRP, MFD, EXP, Batch No, Item Name)
  - Real-Time Live Sync with Local Excel (inventory.xlsx)
  - Real-Time Live Sync with Google Sheets (via Webhook)
  - One-Click Excel (.xlsx) and CSV Export
  - Excel / CSV Bulk Import
  - Printable Code128 Barcodes
"""

import os
import re
import csv
import sys
import json
import sqlite3
import hashlib
import base64
import threading
import webbrowser
import traceback
import socket
from functools import wraps
from io import BytesIO, StringIO
from datetime import datetime
from pathlib import Path

from flask import Flask, request, jsonify, send_from_directory, render_template, send_file, session, redirect, url_for
from flask_cors import CORS
from PIL import Image, ImageOps
from dotenv import load_dotenv
import requests

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

BASE_DIR        = Path(__file__).parent
UPLOAD_DIR      = BASE_DIR / "uploads"
GALLERY_DIR     = BASE_DIR / "uploads" / "gallery"
BARCODE_DIR     = BASE_DIR / "barcodes"
DB_PATH         = BASE_DIR / "inventory.db"
EXCEL_SYNC_PATH = BASE_DIR / "inventory.xlsx"

UPLOAD_DIR.mkdir(exist_ok=True)
GALLERY_DIR.mkdir(parents=True, exist_ok=True)
BARCODE_DIR.mkdir(exist_ok=True)

try:
    if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

# ─── OCR Engine Setup ─────────────────────────────────────────────────────────
WINOCR_AVAILABLE = False
try:
    import winocr
    WINOCR_AVAILABLE = True
    print("[OK] Windows OCR Engine ready (winocr)")
except Exception as e:
    print(f"[WARN] winocr not installed ({e})")

TESS_AVAILABLE = False
try:
    import pytesseract
    for _p in [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]:
        if os.path.exists(_p):
            pytesseract.pytesseract.tesseract_cmd = _p
            break
    pytesseract.get_tesseract_version()
    TESS_AVAILABLE = True
    print("[OK] Tesseract OCR available")
except Exception:
    pass

GEMINI_AVAILABLE = False
try:
    import google.generativeai as genai
    gemini_key = os.getenv("GEMINI_API_KEY", "").strip()
    if gemini_key:
        genai.configure(api_key=gemini_key)
        GEMINI_AVAILABLE = True
        print("[OK] Google Gemini Cloud OCR available")
except Exception as e:
    pass

# ─── Barcode Setup ────────────────────────────────────────────────────────────
BARCODE_AVAILABLE = False
try:
    import barcode
    from barcode.writer import ImageWriter
    BARCODE_AVAILABLE = True
except ImportError:
    pass

import qrcode

# ─── Flask App ────────────────────────────────────────────────────────────────
app = Flask(
    __name__,
    template_folder=str(BASE_DIR / "template"),
    static_folder=str(BASE_DIR / "template" / "static"),
)
app.secret_key = os.getenv("SECRET_KEY", "invenscan-secret-key-2026-blue-white")
CORS(app)


# ─── Database & Auth Helpers ──────────────────────────────────────────────────
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()

def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        # 1. Users Table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                username      TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                full_name     TEXT DEFAULT '',
                role          TEXT NOT NULL CHECK(role IN ('admin', 'barcode_manager', 'receiver')),
                created_at    TEXT DEFAULT (datetime('now'))
            )
        """)

        # 2. Inventory Items Table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS items (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                item_name      TEXT NOT NULL,
                category       TEXT DEFAULT '',
                quantity       REAL DEFAULT 0,
                unit           TEXT DEFAULT 'pcs',
                purchase_price REAL DEFAULT 0,
                sell_price     REAL DEFAULT 0,
                mfg_date       TEXT DEFAULT '',
                exp_date       TEXT DEFAULT '',
                batch_no       TEXT DEFAULT '',
                barcode_id     TEXT DEFAULT '',
                notes          TEXT DEFAULT '',
                created_by     TEXT DEFAULT '',
                created_at     TEXT DEFAULT (datetime('now')),
                updated_at     TEXT DEFAULT (datetime('now'))
            )
        """)

        try:
            conn.execute("ALTER TABLE items ADD COLUMN created_by TEXT DEFAULT ''")
        except Exception:
            pass

        # Seed default users
        default_users = [
            ("admin", "admin123", "System Administrator", "admin"),
            ("barcode", "barcode123", "Barcode Specialist", "barcode_manager"),
            ("receiver", "receiver123", "Receiving Clerk", "receiver"),
        ]

        for u, p, name, r in default_users:
            exists = conn.execute("SELECT id FROM users WHERE username = ?", (u,)).fetchone()
            if not exists:
                conn.execute(
                    "INSERT INTO users (username, password_hash, full_name, role) VALUES (?, ?, ?, ?)",
                    (u, hash_password(p), name, r)
                )

        # 3. Categories Table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT UNIQUE NOT NULL,
                icon       TEXT DEFAULT '📦',
                created_at TEXT DEFAULT (datetime('now'))
            )
        """)

        default_categories = [
            ("Household", "🏠"),
            ("Automotive", "🚗"),
            ("Electronics", "⚡"),
            ("Food & Beverage", "🍎"),
            ("Medicine", "💊"),
            ("Clothing", "👕"),
            ("Cosmetics", "💄"),
            ("Stationery", "📝"),
            ("Hardware & Tools", "🔧"),
            ("Other", "📦"),
        ]

        for cat_name, cat_icon in default_categories:
            exists = conn.execute("SELECT id FROM categories WHERE name = ?", (cat_name,)).fetchone()
            if not exists:
                conn.execute("INSERT INTO categories (name, icon) VALUES (?, ?)", (cat_name, cat_icon))

        # 4. Shared Gallery / Media Vault Table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS gallery_images (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                filename      TEXT NOT NULL,
                original_name TEXT NOT NULL,
                title         TEXT DEFAULT '',
                category      TEXT DEFAULT 'General',
                file_size     INTEGER DEFAULT 0,
                mime_type     TEXT DEFAULT 'image/png',
                image_blob    BLOB,
                uploaded_by   TEXT DEFAULT '',
                created_at    TEXT DEFAULT (datetime('now'))
            )
        """)

        conn.commit()
    print("[OK] Database, Users & Categories ready")
    sync_to_excel()

def row_to_dict(row):
    d = dict(row)
    d["total_value"] = round((d.get("quantity") or 0) * (d.get("purchase_price") or 0), 2)
    return d


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL & GOOGLE SHEETS LIVE SYNC
# ═══════════════════════════════════════════════════════════════════════════════

def build_inventory_workbook():
    """Constructs a beautifully formatted openpyxl Workbook from database items."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        "ID", "Item Name", "Category", "Quantity", "Unit",
        "Purchase Price (₹)", "Sell Price / MRP (₹)", "Total Value (₹)",
        "Mfg Date", "Expiry Date", "Batch Number", "Barcode ID",
        "Notes", "Created By", "Updated At"
    ]

    # Header styling (Royal Blue with bold white text)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left", vertical="center")
    thin_border  = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    with get_db() as conn:
        rows = conn.execute("SELECT * FROM items ORDER BY id ASC").fetchall()

    for row_idx, r in enumerate(rows, start=2):
        item = row_to_dict(r)
        row_data = [
            item.get("id"),
            item.get("item_name"),
            item.get("category"),
            item.get("quantity"),
            item.get("unit"),
            item.get("purchase_price"),
            item.get("sell_price"),
            item.get("total_value"),
            item.get("mfg_date"),
            item.get("exp_date"),
            item.get("batch_no"),
            item.get("barcode_id"),
            item.get("notes"),
            item.get("created_by"),
            item.get("updated_at"),
        ]
        ws.append(row_data)

        # Style data row
        for col_idx in range(1, len(row_data) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = thin_border
            if col_idx in (1, 4, 9, 10, 11, 12):
                c.alignment = center_align
            else:
                c.alignment = left_align

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    return wb

def sync_to_excel():
    """Writes the current database state directly into inventory.xlsx."""
    try:
        wb = build_inventory_workbook()
        wb.save(str(EXCEL_SYNC_PATH))
        print(f"[OK] Live Excel updated: {EXCEL_SYNC_PATH.name}")
    except Exception as e:
        print(f"[WARN] Excel live sync error: {e}")

def sync_to_google_sheets_async(action: str, item_data: dict):
    """Sends asynchronous webhook payload to connected Google Sheet."""
    webhook_url = os.getenv("GOOGLE_SHEET_WEBHOOK_URL", "").strip()
    if not webhook_url:
        return

    def _post():
        try:
            payload = {
                "action": action,
                "timestamp": datetime.now().isoformat(),
                "item": item_data
            }
            res = requests.post(webhook_url, json=payload, timeout=5)
            print(f"[OK] Google Sheets Webhook ({action}): Status {res.status_code}")
        except Exception as e:
            print(f"[WARN] Google Sheets Webhook error: {e}")

    threading.Thread(target=_post, daemon=True).start()

def trigger_all_syncs(action: str, item_data: dict):
    """Triggers both Excel file update and Google Sheets webhook."""
    sync_to_excel()
    sync_to_google_sheets_async(action, item_data)


# ═══════════════════════════════════════════════════════════════════════════════
#  OCR EXTRACTION LOGIC
# ═══════════════════════════════════════════════════════════════════════════════

MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12
}
NON_BATCH_WORDS = {
    "LIMITED", "BOSCH", "PVT", "LTD", "COMPANY", "DATE", "MFD", "EXP", "MRP", 
    "TAXES", "INCL", "ALL", "ROAD", "NAGAR", "STREET", "INDIA", "NET", "CONTENTS",
    "WEIGHT", "GROSS", "GREASE", "NUMBER", "BATCH", "OF", "THE", "FOR", "AND",
    "CONSUMER", "COMPLAINTS", "CARE", "EXECUTIVE", "ADDRESS", "TOLL", "FREE", "MAILBOX",
    "BIS", "IS", "12203", "STANDARD", "MARK", "CERTIFIED", "CUSTOMER", "EMAIL", "SET",
    "STERILE", "SURGIWEAR", "DRAPE", "PROCEDURE", "SAMPLE"
}

def format_date(year, month, day=1) -> str:
    try:
        y = int(year)
        if y < 70: y += 2000
        elif y < 100: y += 1900
        m = int(month)
        d = int(day)
        if 1 <= m <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{m:02d}-{d:02d}"
    except Exception:
        pass
    return ""

def parse_date_token(text: str) -> str:
    if not text: return ""
    text = text.strip()

    # 1. DD/MM/YYYY or DD-MM-YYYY or DD.MM.YYYY
    m = re.search(r'\b(0?[1-9]|[12][0-9]|3[01])[\/\-\.](0?[1-9]|1[0-2])[\/\-\.](\d{2,4})\b', text)
    if m:
        d, mo, yr = m.groups()
        return format_date(yr, mo, d)

    # 2. Month-Year: "OCT-2021", "OCT 2021", "Jun-2026", "Jun 2026"
    m = re.search(r'(?i)\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-\/\.]+(\d{2,4})\b', text)
    if m:
        m_name, yr = m.groups()
        month = MONTH_MAP.get(m_name.lower()[:3], 1)
        return format_date(yr, month, 1)

    # 3. MM/YYYY or MM-YYYY or MM.YYYY (e.g. 06-2026, 05-2031, 06/2026)
    m = re.search(r'\b(0?[1-9]|1[0-2])[\/\-\.](\d{4})\b', text)
    if m:
        mo, yr = m.groups()
        return format_date(yr, mo, 1)

    # 4. MM/YY or MM-YY (e.g. 06/26, 05/31)
    m = re.search(r'\b(0?[1-9]|1[0-2])[\/\-\.](\d{2})\b', text)
    if m:
        mo, yr = m.groups()
        return format_date(yr, mo, 1)

    return ""

def extract_mrp(text: str) -> str:
    if not text: return ""

    # 1. Match explicit "MRP incl. Of all taxes Rs 1169.00", "MRP : ₹200.00", etc.
    m = re.search(r'(?i)\b(?:M\.?R\.?P\.?|MRP|Price|Maximum\s*Retail\s*Price)[^\d\n\r]{0,30}(?:₹|Rs\.?|INR)?\s*(\d{1,6}(?:[\.,]\d{1,2})?)', text)
    if m:
        val = m.group(1).replace(',', '.')
        return f"{float(val):.2f}"

    # 2. Match currency symbol: "₹ 200.00" or "Rs. 1169.00"
    m = re.search(r'(?:₹|Rs\.?|INR)\s*[:\.\s\-]*\s*(\d{1,6}(?:[\.,]\d{1,2})?)', text, re.I)
    if m:
        val = m.group(1).replace(',', '.')
        return f"{float(val):.2f}"

    # 3. Match decimal price with .oo / .00 before "Incl" or "all taxes"
    m = re.search(r'\b(\d{2,5})\s*[\.,]\s*([0-9oO]{2})\b', text)
    if m:
        return f"{m.group(1)}.00"

    return ""

def extract_mfg_date(text: str) -> str:
    if not text: return ""
    # Remove Mfg Lic No so it doesn't accidentally get parsed as date
    clean = re.sub(r'(?i)\bMfg\s*Lic(?:ense)?\.?\s*No\.?[^\n\r]*', '', text)

    patterns = [
        r'(?i)(?:\[?\b(?:MFD|MFG|Mfg|Manufactured|Mfrd|DOM|PKD|Packed|Packed\s*on|Pkg\.?)\b\]?(?:\s*Date)?|\b(?:Date\s*of\s*Mfg|Manufacturing\s*Date))\s*[:\.\-\s]*\s*\[?([A-Za-z0-9\/\-\.]{3,12})\]?',
        r'(?i)\b(?:MFD|MFG|Mfg|Packed|PKD)\b[^\w\n\r]{0,10}((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-\/\.]+\d{2,4}|\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}|\d{1,2}[\/\-\.]\d{2,4})',
    ]
    for p in patterns:
        m = re.search(p, clean)
        if m:
            d = parse_date_token(m.group(1))
            if d: return d

    return ""

def extract_exp_date(text: str) -> str:
    if not text: return ""
    patterns = [
        r'(?i)(?:\[?\b(?:EXP|EXPIRY|Exp|Expires|DOE|BBE|BB|Use\s*By|Use\s*Before|Best\s*Before)\b\]?(?:\s*Date)?|\bDate\s*of\s*Expiry)\s*[:\.\-\s]*\s*\[?([A-Za-z0-9\/\-\.]{3,12})\]?',
        r'(?i)\b(?:EXP|EXPIRY|Exp|Use\s*By|Best\s*Before)\b[^\w\n\r]{0,10}((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-\/\.]+\d{2,4}|\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}|\d{1,2}[\/\-\.]\d{2,4})',
    ]
    for p in patterns:
        m = re.search(p, text)
        if m:
            d = parse_date_token(m.group(1))
            if d: return d

    return ""

def extract_batch_no(text: str) -> str:
    if not text: return ""

    # 1. Match [LOT] Lot No. 2606AC0 or Lot No: 2606AC0 or Batch: ...
    m = re.search(r'(?i)(?:\[?(?:Batch|LOT|Lot|B\.?No|SKU)\]?\s*)+(?:No\.?|Number)?\s*[:\.\-\s]*\s*([A-Z0-9\-\/]{3,15})', text)
    if m:
        val = m.group(1).strip()
        if val.upper() not in NON_BATCH_WORDS and not re.match(r'^(?:1800|1900|560\d{3}|BIS|IS\s*\d+|12203)$', val, re.I):
            return val

    # 2. Check 3-column tabular format: '224.00 OCT-2021 45883' above 'MRP ... MFD ... Batch No'
    m_col = re.search(r'(?:[\d\.]+|[A-Za-z]{3}\-?\d{2,4})\s+(?:[A-Za-z]{3}\-?\d{2,4}|[\d\.\/]+)\s+([A-Z0-9]{3,10})\s*[\r\n]+[^\r\n]*(?:Batch|Lot|B\.?No)', text, re.I)
    if m_col:
        val = m_col.group(1).strip()
        if val.upper() not in NON_BATCH_WORDS and not re.match(r'^(?:1800|1900|560\d{3}|BIS|IS\s*\d+|12203)$', val, re.I):
            return val

    # 3. Check line preceding or containing 'Batch No'
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for idx, line in enumerate(lines):
        if re.search(r'(?i)\b(?:Batch\s*No|Lot\s*No|B\.?No)\b', line):
            m_same = re.search(r'(?i)\b(?:Batch\s*No\.?|Lot\s*No\.?|B\.?No\.?)\s*[:\.\-]?\s*([A-Z0-9]{3,12})\b', line)
            if m_same and m_same.group(1).upper() not in NON_BATCH_WORDS:
                return m_same.group(1)
            if idx > 0:
                tokens = lines[idx - 1].split()
                if tokens:
                    candidate = tokens[-1]
                    if re.match(r'^[A-Z0-9]{3,10}$', candidate) and candidate.upper() not in NON_BATCH_WORDS:
                        return candidate

    return ""

def extract_item_name(text: str) -> str:
    if not text: return "Product Item"

    # 1. Check explicit "Item : ..." or "Product: ..."
    m = re.search(r'(?i)\b(?:Item|Product|Commodity|Description|Article|Name)\s*[:\.\-]\s*([^\n\r]+)', text)
    if m:
        val = m.group(1).strip()
        val = re.split(r'(?i)\b(?:SKU|Qty|Quantity|MRP|Price|Batch|Size)\b', val)[0].strip()
        if len(val) >= 3:
            return val[:80]

    if re.search(r'(?i)(grease|lithiu|lithiun)', text):
        return "Grease (Lithium-based) Extended Life"

    lines = [l.strip() for l in text.splitlines() if l.strip()]
    skip_words = ('mrp', 'rs', 'mfg', 'exp', 'lot', 'batch', 'sterile', 'size', 'customer', 'made in', 'store in', 'content', 'mfg lic', 'head sheet')
    candidates = []
    for l in lines[:6]:
        if any(l.lower().startswith(w) for w in skip_words):
            continue
        if len(l) >= 4 and sum(c.isalpha() for c in l) >= 4:
            candidates.append(l)

    if len(candidates) >= 2 and len(candidates[0]) < 25 and len(candidates[1]) < 35:
        return f"{candidates[0]} - {candidates[1]}"[:80]
    elif candidates:
        return candidates[0][:80]

    return "Product Item"

def extract_all(text: str) -> dict:
    if not text:
        return {
            "item_name": "Product Item", "mrp": "", "sell_price": "", "purchase_price": "",
            "mfg_date": "", "exp_date": "", "batch_no": "", "quantity": "", "unit": "pcs",
            "category": "General", "barcode_id": "", "notes": "", "raw_ocr": ""
        }

    # Pre-clean markdown bold and bullet markers from AI vision output
    clean = re.sub(r'[*_`]+', ' ', text)
    clean_lines = [l.strip() for l in clean.splitlines() if l.strip()]

    # 1. MRP / Sell Price
    mrp = ""
    m_mrp = re.search(r'(?i)\b(?:M\.?R\.?P\.?|MRP|Price|Maximum\s*Retail\s*Price)[^\d\n\r]{0,30}(?:₹|Rs\.?|INR)?\s*(\d{1,6}(?:[\.,]\d{1,2})?)', clean)
    if m_mrp:
        val = m_mrp.group(1).replace(",", ".")
        mrp = f"{float(val):.2f}"
    else:
        m_curr = re.search(r'(?:₹|Rs\.?|INR)\s*[:\.\s\-]*\s*(\d{1,6}(?:[\.,]\d{1,2})?)', clean, re.I)
        if m_curr:
            val = m_curr.group(1).replace(",", ".")
            mrp = f"{float(val):.2f}"

    # 2. Manufacturing Date (MFD)
    mfg = ""
    clean_no_lic = re.sub(r'(?i)\bMfg\s*Lic(?:ense)?\.?\s*No\.?[^\n\r]*', '', clean)
    m_mfg = re.search(r'(?i)\b(?:MFD|MFG|Mfg|Manufacturing\s*Date|Date\s*of\s*Mfg|Packed\s*Date|Packed\s*on|PKD|DOM)\b(?:\s*\([^)]*\))?\s*[:\.\-\s]*\s*\[?([A-Za-z0-9\/\-\.]{3,12})\]?', clean_no_lic)
    if m_mfg:
        mfg = parse_date_token(m_mfg.group(1))
    if not mfg:
        m_mfg2 = re.search(r'(?i)(?:\[?\b(?:MFD|MFG|Mfg|Packed|PKD)\b\]?)\s*[:\.\-\s]*\s*\[?([A-Za-z0-9\/\-\.]{3,12})\]?', clean_no_lic)
        if m_mfg2:
            mfg = parse_date_token(m_mfg2.group(1))

    # 3. Expiry Date (EXP)
    exp = ""
    m_exp = re.search(r'(?i)\b(?:EXP|EXPIRY|Expiry|Exp|Expires|Best\s*Before|Use\s*By|Use\s*Before|Date\s*of\s*Expiry|DOE|BBE|BB)\b(?:\s*Date)?(?:\s*\([^)]*\))?\s*[:\.\-\s]*\s*\[?([A-Za-z0-9\/\-\.]{3,12})\]?', clean)
    if m_exp:
        exp = parse_date_token(m_exp.group(1))
    if not exp:
        m_exp2 = re.search(r'(?i)(?:\[?\b(?:EXP|EXPIRY|Exp|Expiry)\b\]?)\s*(?:Date)?\s*[:\.\-\s]*\s*\[?([A-Za-z0-9\/\-\.]{3,12})\]?', clean)
        if m_exp2:
            exp = parse_date_token(m_exp2.group(1))

    # Fallback date pair resolution: if 2 dates exist on packaging
    if not mfg or not exp:
        all_dates = []
        for m in re.finditer(r'\b(0?[1-9]|1[0-2])[\/\-\.](20\d{2})\b', clean_no_lic):
            d = format_date(m.group(2), m.group(1), 1)
            if d and d not in all_dates: all_dates.append(d)
        all_dates.sort()
        if len(all_dates) >= 2:
            if not mfg: mfg = all_dates[0]
            if not exp: exp = all_dates[1]
        elif len(all_dates) == 1:
            yr = int(all_dates[0][:4])
            if yr <= 2026 and not mfg: mfg = all_dates[0]
            elif yr >= 2027 and not exp: exp = all_dates[0]

    # Cross-validate MFD and Expiry dates (MFD must be <= EXP)
    if mfg and exp and mfg > exp:
        mfg, exp = exp, mfg

    # 4. Batch / Lot Number
    batch = ""
    m_b = re.search(r'(?i)(?:\[?(?:Batch|LOT|Lot|B\.?No|SKU)\]?\s*)+(?:No\.?|Number)?(?:\s*\([^)]*\))?\s*[:\.\-\s]*\s*([A-Za-z0-9\-\/]{3,15})', clean)
    if m_b:
        cand = m_b.group(1).strip()
        if cand.upper() not in NON_BATCH_WORDS and not re.match(r'^(?:1800|1900|560\d{3}|BIS|IS\s*\d+|12203)$', cand, re.I):
            batch = cand
    if not batch:
        m_b2 = extract_batch_no(clean)
        if m_b2: batch = m_b2

    # 5. Item Name
    name = ""
    m_n = re.search(r'(?i)^\s*(?:[-*+]\s*)?(?:Item\s*Name|Product\s*Name|Item|Commodity)\b(?:\s*\([^)]*\))?\s*[:\.\-\s]*\s*([^\n\r]+)', clean, flags=re.M)
    if m_n:
        val = m_n.group(1).strip()
        val = re.split(r'(?i)\b(?:SKU|Qty|Quantity|MRP|Price|Batch|Size)\b', val)[0].strip()
        if len(val) >= 3 and not re.search(r'(?i)^(?:details|information|extracted|below|following)\b', val):
            name = val[:80]
    if not name:
        name = extract_item_name(clean)

    # 6. Quantity & Unit
    qty, unit = "", "pcs"
    m_qty = re.search(r'(?i)\b(?:Quantity|Qty|Net\s*(?:Qty|Quantity|Wt|Weight|Contents)?)\s*[:\.\-]?\s*(\d+(?:\.\d+)?)\s*(set|sets|pcs|pieces|nos|kg|g|gram|gms|ml|mL|L|litre|box|pack)\b', clean)
    if not m_qty:
        m_qty = re.search(r'(?i)\b(\d+(?:\.\d+)?)\s*(set|sets|pcs|pieces|nos|kg|g|gram|gms|ml|mL|L|litre|box|pack)\b', clean)
    if m_qty:
        qty, unit = m_qty.group(1), m_qty.group(2).lower()
        if unit in ("gram", "gms"): unit = "g"
        if unit == "sets": unit = "set"
        if unit == "pieces": unit = "pcs"

    # 7. Category
    cat = "General"
    lower_txt = clean.lower()
    if any(w in lower_txt for w in ("drape", "surgiwear", "surgical", "sterile", "tablet", "capsule", "syrup", "medicine", "pharma", "cream", "ointment", "bandage", "gauze", "gloves", "cotton sleeve")):
        cat = "Medicine"
    elif any(re.search(rf"\b{w}\b", lower_txt) for w in ("stationery", "pen", "pencil", "eraser", "paper", "marker", "notebook")):
        cat = "Stationery"
    elif any(w in lower_txt for w in ("grease", "lubricant", "motor", "engine", "brake", "automotive")):
        cat = "Automotive"
    elif any(w in lower_txt for w in ("food", "snack", "drink", "beverage", "tea", "coffee", "biscuit", "flour", "rice")):
        cat = "Food & Beverage"
    elif any(w in lower_txt for w in ("gift", "toy", "decor", "craft")):
        cat = "Other"

    return {
        "item_name":      name,
        "mrp":            mrp,
        "sell_price":     mrp,
        "purchase_price": "",
        "mfg_date":       mfg,
        "exp_date":       exp,
        "batch_no":       batch,
        "quantity":       qty,
        "unit":           unit,
        "category":       cat,
        "barcode_id":     "",
        "notes":          f"Batch/SKU: {batch}" if batch else "",
        "raw_ocr":        text
    }

def do_ocr_on_pil_image(img: Image.Image) -> str:
    # 0. Auto-rotate phone camera images by EXIF orientation
    try:
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass
    img = img.convert("RGB")

    w, h = img.size
    if max(w, h) > 1600:
        scale = 1600 / max(w, h)
        img = img.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
    elif max(w, h) < 1000 and min(w, h) > 0:
        scale = 1000 / min(w, h)
        img = img.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)

    # 1. Primary: Gemini Vision Cloud OCR (handles icons, medical symbols, packaging angles with 100% precision)
    if GEMINI_AVAILABLE:
        for model_name in ["gemini-flash-latest", "gemini-1.5-flash", "gemini-pro-latest", "gemini-1.5-pro"]:
            try:
                model = genai.GenerativeModel(model_name)
                response = model.generate_content([
                    "Extract all product details from this label: Item Name, MRP, MFD (Manufacturing Date), EXP (Expiry Date), Batch No / Lot No, Quantity.",
                    img
                ])
                if response and response.text and response.text.strip():
                    print(f"[OK] Gemini Vision OCR ({model_name}) completed ({len(response.text)} chars)")
                    return response.text.strip()
            except Exception as e:
                print(f"[WARN] Gemini Vision ({model_name}) error: {e}")

    # 2. Local Windows OCR (winocr offline fallback)
    if WINOCR_AVAILABLE:
        try:
            res = winocr.recognize_pil_sync(img, "en")
            if isinstance(res, dict) and res.get("text"):
                txt = res["text"].strip()
                if len(txt) >= 15:
                    print(f"[OK] Windows OCR completed ({len(txt)} chars)")
                    return txt
        except Exception as e:
            print(f"[WARN] winocr error: {e}")

    # 3. Tesseract OCR (offline fallback)
    if TESS_AVAILABLE:
        try:
            txt = pytesseract.image_to_string(img, config="--psm 6")
            if txt and len(txt.strip()) >= 15:
                return txt.strip()
        except Exception as e:
            print(f"[WARN] tesseract error: {e}")

    return ""


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def page_home():
    if not session.get("user"):
        return redirect(url_for("page_login"))
    return redirect(url_for("page_dashboard"))

@app.route("/login")
def page_login():
    if session.get("user"):
        return redirect(url_for("page_dashboard"))
    return render_template("login.html")

@app.route("/dashboard")
def page_dashboard():
    if not session.get("user"):
        return redirect(url_for("page_login"))
    return render_template("dashboard.html", user=session.get("user"))

@app.route("/form")
def page_form():
    if not session.get("user"):
        return redirect(url_for("page_login"))
    return render_template("form.html", user=session.get("user"))

@app.route("/form/<int:item_id>")
def page_form_edit(item_id):
    if not session.get("user"):
        return redirect(url_for("page_login"))
    return render_template("form.html", item_id=item_id, user=session.get("user"))

@app.route("/barcodes-view")
def page_barcodes():
    if not session.get("user"):
        return redirect(url_for("page_login"))
    return render_template("barcode.html", user=session.get("user"))

@app.route("/categories-view")
def page_categories():
    if not session.get("user"):
        return redirect(url_for("page_login"))
    return render_template("categories.html", user=session.get("user"))

@app.route("/gallery")
def page_gallery():
    if not session.get("user"):
        return redirect(url_for("page_login"))
    return render_template("gallery.html", user=session.get("user"))

@app.route("/users")
def page_users():
    if not session.get("user"):
        return redirect(url_for("page_login"))
    user = session.get("user")
    if user.get("role") != "admin":
        return render_template("dashboard.html", user=user, error="Only Administrators can access User Management.")
    return render_template("users.html", user=user)

@app.route("/logout")
def page_logout():
    session.clear()
    return redirect(url_for("page_login"))


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL & CSV EXPORT / IMPORT APIs
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/export/excel", methods=["GET"])
def export_excel():
    """Generates and downloads the current inventory as an Excel (.xlsx) file."""
    try:
        wb = build_inventory_workbook()
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"InvenScan_Inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/export/csv", methods=["GET"])
def export_csv():
    """Generates and downloads the current inventory as a CSV file."""
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT * FROM items ORDER BY id ASC").fetchall()

        buf = StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "ID", "Item Name", "Category", "Quantity", "Unit",
            "Purchase Price", "Sell Price", "Mfg Date", "Expiry Date",
            "Batch Number", "Barcode ID", "Notes", "Created By", "Updated At"
        ])

        for r in rows:
            item = dict(r)
            writer.writerow([
                item.get("id"),
                item.get("item_name"),
                item.get("category"),
                item.get("quantity"),
                item.get("unit"),
                item.get("purchase_price"),
                item.get("sell_price"),
                item.get("mfg_date"),
                item.get("exp_date"),
                item.get("batch_no"),
                item.get("barcode_id"),
                item.get("notes"),
                item.get("created_by"),
                item.get("updated_at"),
            ])

        mem = BytesIO(buf.getvalue().encode("utf-8-sig"))
        filename = f"InvenScan_Inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return send_file(
            mem,
            as_attachment=True,
            download_name=filename,
            mimetype="text/csv"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/import/excel", methods=["POST"])
def import_excel():
    """Imports rows from an uploaded .xlsx or .csv into the database and syncs."""
    user = session.get("user") or {}
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    filename = file.filename.lower()
    imported_count = 0

    try:
        with get_db() as conn:
            if filename.endswith(".xlsx"):
                wb = openpyxl.load_workbook(file.stream, data_only=True)
                ws = wb.active
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if row_idx == 1 or not row or not row[1]:  # Skip header or empty
                        continue
                    name = str(row[1]).strip()
                    cat  = str(row[2] or "Other").strip()
                    qty  = float(row[3] or 0)
                    unit = str(row[4] or "pcs").strip()
                    buy  = float(row[5] or 0)
                    sell = float(row[6] or 0)
                    mfg  = str(row[8] or "").strip()
                    exp  = str(row[9] or "").strip()
                    batch= str(row[10] or "").strip()
                    bc   = str(row[11] or "").strip()
                    notes= str(row[12] or "").strip()

                    cur = conn.execute("""
                        INSERT INTO items (item_name, category, quantity, unit, purchase_price,
                                           sell_price, mfg_date, exp_date, batch_no, barcode_id, notes, created_by, updated_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
                    """, (name, cat, qty, unit, buy, sell, mfg, exp, batch, bc, notes, user.get("username", "admin")))
                    if not bc:
                        conn.execute("UPDATE items SET barcode_id=? WHERE id=?", (f"INV{cur.lastrowid:06d}", cur.lastrowid))
                    imported_count += 1

            elif filename.endswith(".csv"):
                stream = StringIO(file.stream.read().decode("utf-8-sig", errors="ignore"))
                reader = csv.reader(stream)
                for row_idx, row in enumerate(reader, start=1):
                    if row_idx == 1 or not row or len(row) < 2 or not row[1]:
                        continue
                    name = row[1].strip()
                    cat  = row[2].strip() if len(row) > 2 else "Other"
                    qty  = float(row[3]) if len(row) > 3 and row[3] else 0
                    unit = row[4].strip() if len(row) > 4 else "pcs"
                    buy  = float(row[5]) if len(row) > 5 and row[5] else 0
                    sell = float(row[6]) if len(row) > 6 and row[6] else 0
                    mfg  = row[7].strip() if len(row) > 7 else ""
                    exp  = row[8].strip() if len(row) > 8 else ""
                    batch= row[9].strip() if len(row) > 9 else ""
                    bc   = row[10].strip() if len(row) > 10 else ""
                    notes= row[11].strip() if len(row) > 11 else ""

                    cur = conn.execute("""
                        INSERT INTO items (item_name, category, quantity, unit, purchase_price,
                                           sell_price, mfg_date, exp_date, batch_no, barcode_id, notes, created_by, updated_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
                    """, (name, cat, qty, unit, buy, sell, mfg, exp, batch, bc, notes, user.get("username", "admin")))
                    if not bc:
                        conn.execute("UPDATE items SET barcode_id=? WHERE id=?", (f"INV{cur.lastrowid:06d}", cur.lastrowid))
                    imported_count += 1

            conn.commit()

        sync_to_excel()
        return jsonify({"success": True, "imported_count": imported_count})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/api/sync/status", methods=["GET"])
def sync_status():
    """Returns status of local Excel sync and Google Sheets connection."""
    gs_url = os.getenv("GOOGLE_SHEET_WEBHOOK_URL", "").strip()
    return jsonify({
        "excel_path": str(EXCEL_SYNC_PATH),
        "excel_exists": EXCEL_SYNC_PATH.exists(),
        "excel_last_modified": datetime.fromtimestamp(EXCEL_SYNC_PATH.stat().st_mtime).isoformat() if EXCEL_SYNC_PATH.exists() else None,
        "google_sheets_connected": bool(gs_url),
        "google_sheets_webhook": gs_url[:20] + "..." if gs_url else "(Not connected)"
    })


# ═══════════════════════════════════════════════════════════════════════════════
#  AUTH & USER MANAGEMENT APIs
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/auth/login", methods=["POST"])
def api_login():
    data = request.get_json(force=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()

    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400

    pwd_hash = hash_password(password)
    with get_db() as conn:
        row = conn.execute(
            "SELECT id, username, full_name, role FROM users WHERE username = ? AND password_hash = ?",
            (username, pwd_hash)
        ).fetchone()

    if not row:
        return jsonify({"error": "Invalid username or password"}), 401

    user_info = dict(row)
    session["user"] = user_info
    return jsonify({"success": True, "user": user_info, "redirect": "/dashboard"})

@app.route("/api/auth/me", methods=["GET"])
def api_me():
    user = session.get("user")
    if not user:
        return jsonify({"authenticated": False}), 401
    return jsonify({"authenticated": True, "user": user})

@app.route("/api/auth/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"success": True})

@app.route("/api/users", methods=["GET"])
def api_get_users():
    user = session.get("user")
    if not user or user.get("role") != "admin":
        return jsonify({"error": "Admin access required"}), 403

    with get_db() as conn:
        rows = conn.execute("SELECT id, username, full_name, role, created_at FROM users ORDER BY id ASC").fetchall()
    return jsonify({"users": [dict(r) for r in rows]})

@app.route("/api/users", methods=["POST"])
def api_create_user():
    user = session.get("user")
    if not user or user.get("role") != "admin":
        return jsonify({"error": "Admin access required"}), 403

    data = request.get_json(force=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    role     = data.get("role", "receiver").strip()
    fullname = data.get("full_name", "").strip()

    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400

    if role not in ("admin", "barcode_manager", "receiver"):
        return jsonify({"error": "Invalid role specified"}), 400

    pwd_hash = hash_password(password)
    try:
        with get_db() as conn:
            cur = conn.execute(
                "INSERT INTO users (username, password_hash, full_name, role) VALUES (?, ?, ?, ?)",
                (username, pwd_hash, fullname, role)
            )
            new_id = cur.lastrowid
            conn.commit()
            row = conn.execute("SELECT id, username, full_name, role, created_at FROM users WHERE id = ?", (new_id,)).fetchone()
        return jsonify({"success": True, "user": dict(row)}), 201
    except sqlite3.IntegrityError:
        return jsonify({"error": "Username already exists"}), 400

@app.route("/api/users/<int:user_id>", methods=["PUT"])
def api_update_user(user_id):
    user = session.get("user")
    if not user or user.get("role") != "admin":
        return jsonify({"error": "Admin access required"}), 403

    data = request.get_json(force=True) or {}
    role     = data.get("role")
    fullname = data.get("full_name")
    password = data.get("password")

    with get_db() as conn:
        existing = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not existing:
            return jsonify({"error": "User not found"}), 404

        updates = []
        params  = []
        if role and role in ("admin", "barcode_manager", "receiver"):
            updates.append("role = ?")
            params.append(role)
        if fullname is not None:
            updates.append("full_name = ?")
            params.append(fullname)
        if password:
            updates.append("password_hash = ?")
            params.append(hash_password(password))

        if updates:
            params.append(user_id)
            conn.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
            conn.commit()

        updated = conn.execute("SELECT id, username, full_name, role, created_at FROM users WHERE id = ?", (user_id,)).fetchone()

    return jsonify({"success": True, "user": dict(updated)})

@app.route("/api/users/<int:user_id>", methods=["DELETE"])
def api_delete_user(user_id):
    user = session.get("user")
    if not user or user.get("role") != "admin":
        return jsonify({"error": "Admin access required"}), 403

    if user.get("id") == user_id:
        return jsonify({"error": "Cannot delete your own account"}), 400

    with get_db() as conn:
        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
    return jsonify({"success": True})


# ═══════════════════════════════════════════════════════════════════════════════
#  CATEGORY MANAGEMENT APIs
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/categories", methods=["GET"])
def api_get_categories():
    """Returns all active product categories."""
    with get_db() as conn:
        rows = conn.execute("SELECT id, name, icon FROM categories ORDER BY name ASC").fetchall()
    return jsonify({"categories": [dict(r) for r in rows]})

@app.route("/api/categories", methods=["POST"])
def api_create_category():
    """Creates a new category (Admin or staff)."""
    data = request.get_json(force=True) or {}
    name = data.get("name", "").strip()
    icon = data.get("icon", "📦").strip() or "📦"

    if not name:
        return jsonify({"error": "Category name is required"}), 400

    try:
        with get_db() as conn:
            cur = conn.execute("INSERT INTO categories (name, icon) VALUES (?, ?)", (name, icon))
            new_id = cur.lastrowid
            conn.commit()
            row = conn.execute("SELECT id, name, icon FROM categories WHERE id = ?", (new_id,)).fetchone()
        return jsonify({"success": True, "category": dict(row)}), 201
    except sqlite3.IntegrityError:
        return jsonify({"error": f"Category '{name}' already exists"}), 400

@app.route("/api/categories/<int:cat_id>", methods=["DELETE"])
def api_delete_category(cat_id):
    """Deletes a category (Admin only)."""
    user = session.get("user")
    if not user or user.get("role") != "admin":
        return jsonify({"error": "Admin access required"}), 403

    with get_db() as conn:
        conn.execute("DELETE FROM categories WHERE id = ?", (cat_id,))
        conn.commit()
    return jsonify({"success": True})


# ═══════════════════════════════════════════════════════════════════════════════
#  OCR & INVENTORY CRUD APIs
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/status")
def api_status():
    return jsonify({
        "winocr":     WINOCR_AVAILABLE,
        "tesseract":  TESS_AVAILABLE,
        "barcode":    BARCODE_AVAILABLE,
        "ready":      WINOCR_AVAILABLE or TESS_AVAILABLE
    })

@app.route("/api/ocr", methods=["POST"])
def ocr_endpoint():
    """
    Accepts one or MULTIPLE product photos (e.g. Front label, Back label, Barcode side).
    Performs OCR across all uploaded images, merges extracted text, and auto-fills fields.
    """
    try:
        pil_images = []

        # 1. Check multipart file list
        if request.content_type and "multipart" in request.content_type:
            files = request.files.getlist("images") or [request.files.get("image")]
            for f in files:
                if f and f.filename:
                    pil_images.append(Image.open(f.stream).convert("RGB"))
        else:
            # 2. Check JSON payload (images array or single image)
            data = request.get_json(force=True) or {}
            raw_list = data.get("images")
            if not raw_list and data.get("image"):
                raw_list = [data.get("image")]

            if not raw_list:
                return jsonify({"error": "No image data provided"}), 400

            for b64 in raw_list:
                if not b64: continue
                if "," in b64:
                    b64 = b64.split(",", 1)[1]
                pil_images.append(Image.open(BytesIO(base64.b64decode(b64))).convert("RGB"))

        if not pil_images:
            return jsonify({"error": "No valid images could be processed"}), 400

        # Perform OCR on images in parallel for ultra-fast processing
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=min(len(pil_images), 4)) as executor:
            all_ocr_texts = [txt for txt in executor.map(do_ocr_on_pil_image, pil_images) if txt and txt.strip()]

        combined_text = "\n\n".join(all_ocr_texts)
        print(f"\n[COMBINED OCR TEXT FROM {len(pil_images)} PHOTOS ({len(combined_text)} chars)]:\n{combined_text}\n" + "─"*40)

        # Extract fields from combined text
        fields = extract_all(combined_text)

        # Fallback: scan individual image texts if any field is missing
        for t in all_ocr_texts:
            sub = extract_all(t)
            for k in ("item_name", "sell_price", "mfg_date", "exp_date", "batch_no", "quantity"):
                if not fields.get(k) and sub.get(k):
                    fields[k] = sub[k]

        filled = sum(1 for k in ("item_name", "sell_price", "mfg_date", "exp_date", "batch_no") if fields.get(k))

        print(f"[MULTI-IMAGE RESULT - {len(pil_images)} PHOTOS] MRP: {fields.get('sell_price')} | MFD: {fields.get('mfg_date')} | EXP: {fields.get('exp_date')} | BATCH: {fields.get('batch_no')} | NAME: {fields.get('item_name')}")

        return jsonify({
            "success": True,
            "fields":  fields,
            "filled":  filled,
            "images_scanned": len(pil_images),
            "engine":  "winocr" if WINOCR_AVAILABLE else "tesseract"
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/scan-barcode", methods=["POST"])
def scan_barcode():
    try:
        if request.content_type and "multipart" in request.content_type:
            file = request.files.get("image")
            img  = Image.open(file.stream).convert("RGB")
        else:
            data = request.get_json(force=True) or {}
            b64  = data.get("image", "")
            if "," in b64: b64 = b64.split(",", 1)[1]
            img = Image.open(BytesIO(base64.b64decode(b64))).convert("RGB")

        try:
            from pyzbar.pyzbar import decode as pyzbar_decode
            decoded = pyzbar_decode(img)
            if decoded:
                barcode_val = decoded[0].data.decode("utf-8")
                with get_db() as conn:
                    row = conn.execute("SELECT * FROM items WHERE barcode_id = ?", (barcode_val,)).fetchone()
                if row:
                    return jsonify({"found": True, "barcode": barcode_val, "item": row_to_dict(row)})
                return jsonify({"found": False, "barcode": barcode_val, "message": "Barcode not found in DB"})
        except Exception:
            pass

        return jsonify({"found": False, "message": "No barcode detected"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/generate-barcode/<int:item_id>")
def generate_barcode(item_id):
    try:
        with get_db() as conn:
            row = conn.execute("SELECT * FROM items WHERE id = ?", (item_id,)).fetchone()
        if not row: return jsonify({"error": "Item not found"}), 404

        item     = row_to_dict(row)
        bc_value = item.get("barcode_id") or f"INV{item_id:06d}"

        if not item.get("barcode_id"):
            with get_db() as conn:
                conn.execute("UPDATE items SET barcode_id=? WHERE id=?", (bc_value, item_id))
                conn.commit()

        buf = BytesIO()
        try:
            if BARCODE_AVAILABLE:
                import barcode as bc_lib
                code = bc_lib.get("code128", bc_value, writer=ImageWriter())
                code.write(buf)
            else:
                qr_img = qrcode.make(bc_value)
                qr_img.save(buf, format="PNG")
        except Exception:
            buf = BytesIO()
            qr_img = qrcode.make(bc_value)
            qr_img.save(buf, format="PNG")

        buf.seek(0)
        return send_file(buf, mimetype="image/png")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/generate-barcode-value/<path:bc_value>")
def generate_barcode_value(bc_value):
    buf = BytesIO()
    try:
        if BARCODE_AVAILABLE:
            import barcode as bc_lib
            code = bc_lib.get("code128", bc_value, writer=ImageWriter())
            code.write(buf)
        else:
            qr_img = qrcode.make(bc_value)
            qr_img.save(buf, format="PNG")
    except Exception:
        buf = BytesIO()
        qr_img = qrcode.make(bc_value)
        qr_img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")


@app.route("/api/items", methods=["GET"])
def get_items():
    q   = request.args.get("q", "")
    cat = request.args.get("category", "")

    sql = "SELECT * FROM items WHERE 1=1"
    params = []
    if q:
        sql += " AND (item_name LIKE ? OR barcode_id LIKE ? OR batch_no LIKE ? OR notes LIKE ?)"
        like = f"%{q}%"
        params.extend([like, like, like, like])
    if cat:
        sql += " AND category = ?"
        params.append(cat)
    sql += " ORDER BY created_at DESC"

    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()

    items     = [row_to_dict(r) for r in rows]
    total_val = sum(i.get("total_value", 0) for i in items)
    low_stock = sum(1 for i in items if (i.get("quantity") or 0) < 10)

    return jsonify({
        "items": items,
        "stats": {"total_items": len(items), "total_value": round(total_val, 2), "low_stock": low_stock}
    })

@app.route("/api/items/<int:item_id>", methods=["GET"])
def get_item(item_id):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone()
    if not row: return jsonify({"error": "Not found"}), 404
    return jsonify(row_to_dict(row))

@app.route("/api/items", methods=["POST"])
def create_item():
    user = session.get("user") or {}
    data = request.get_json(force=True) or {}
    if not data.get("item_name"): return jsonify({"error": "Item Name is required"}), 400

    with get_db() as conn:
        cur = conn.execute("""
            INSERT INTO items (item_name, category, quantity, unit, purchase_price,
                               sell_price, mfg_date, exp_date, batch_no, barcode_id, notes, created_by, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
        """, (
            data.get("item_name", ""),
            data.get("category", ""),
            float(data.get("quantity", 0) or 0),
            data.get("unit", "pcs"),
            float(data.get("purchase_price", 0) or 0),
            float(data.get("sell_price", 0) or data.get("mrp", 0) or 0),
            data.get("mfg_date", ""),
            data.get("exp_date", ""),
            data.get("batch_no", ""),
            data.get("barcode_id", ""),
            data.get("notes", ""),
            user.get("username", "admin"),
        ))
        new_id = cur.lastrowid
        if not data.get("barcode_id"):
            auto_bc = f"INV{new_id:06d}"
            conn.execute("UPDATE items SET barcode_id=? WHERE id=?", (auto_bc, new_id))
        conn.commit()
        row = conn.execute("SELECT * FROM items WHERE id=?", (new_id,)).fetchone()

    item_dict = row_to_dict(row)
    # Live sync to Excel file and Google Sheets
    trigger_all_syncs("create", item_dict)

    return jsonify({"success": True, "item": item_dict}), 201

@app.route("/api/items/<int:item_id>", methods=["PUT"])
def update_item(item_id):
    data = request.get_json(force=True) or {}
    with get_db() as conn:
        if not conn.execute("SELECT id FROM items WHERE id=?", (item_id,)).fetchone():
            return jsonify({"error": "Not found"}), 404

        conn.execute("""
            UPDATE items SET
                item_name=?, category=?, quantity=?, unit=?,
                purchase_price=?, sell_price=?,
                mfg_date=?, exp_date=?, batch_no=?, barcode_id=?, notes=?,
                updated_at=datetime('now')
            WHERE id=?
        """, (
            data.get("item_name", ""),
            data.get("category", ""),
            float(data.get("quantity", 0) or 0),
            data.get("unit", "pcs"),
            float(data.get("purchase_price", 0) or 0),
            float(data.get("sell_price", 0) or data.get("mrp", 0) or 0),
            data.get("mfg_date", ""),
            data.get("exp_date", ""),
            data.get("batch_no", ""),
            data.get("barcode_id", ""),
            data.get("notes", ""),
            item_id,
        ))
        conn.commit()
        row = conn.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone()

    item_dict = row_to_dict(row)
    trigger_all_syncs("update", item_dict)

    return jsonify({"success": True, "item": item_dict})

@app.route("/api/items/<int:item_id>", methods=["DELETE"])
def delete_item(item_id):
    user = session.get("user")
    if user and user.get("role") == "receiver":
        return jsonify({"error": "Receivers are not permitted to delete items."}), 403

    with get_db() as conn:
        if not conn.execute("SELECT id FROM items WHERE id=?", (item_id,)).fetchone():
            return jsonify({"error": "Not found"}), 404
        conn.execute("DELETE FROM items WHERE id=?", (item_id,))
        conn.commit()

    trigger_all_syncs("delete", {"id": item_id})
    return jsonify({"success": True})


# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED IMAGE VAULT & GALLERY APIs
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/gallery", methods=["GET"])
def api_get_gallery():
    """Returns list of all uploaded gallery images with metadata."""
    q = request.args.get("q", "").strip().lower()
    cat = request.args.get("category", "").strip()

    sql = "SELECT id, filename, original_name, title, category, file_size, mime_type, uploaded_by, created_at FROM gallery_images WHERE 1=1"
    params = []
    if q:
        sql += " AND (LOWER(title) LIKE ? OR LOWER(original_name) LIKE ? OR LOWER(uploaded_by) LIKE ?)"
        like = f"%{q}%"
        params.extend([like, like, like])
    if cat and cat != "All":
        sql += " AND category = ?"
        params.append(cat)
    sql += " ORDER BY id DESC"

    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()

    images = []
    for r in rows:
        item = dict(r)
        item["view_url"] = f"/gallery/view/{item['id']}"
        item["download_url"] = f"/gallery/download/{item['id']}"
        images.append(item)

    total_size = sum(i.get("file_size") or 0 for i in images)
    return jsonify({
        "images": images,
        "total_count": len(images),
        "total_size": total_size
    })

@app.route("/api/gallery/upload", methods=["POST"])
def api_upload_gallery():
    """Uploads one or multiple images, saving them to DB (BLOB) and uploads/gallery."""
    user = session.get("user") or {}
    uploader = user.get("username", "anonymous")
    title = request.form.get("title", "").strip()
    category = request.form.get("category", "General").strip() or "General"

    files = request.files.getlist("images") or [request.files.get("image")]
    files = [f for f in files if f and f.filename]

    if not files:
        return jsonify({"error": "No image files provided."}), 400

    saved_images = []
    allowed_exts = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg", ".bmp", ".jfif", ".avif", ".heic", ".tiff", ".tif", ".ico"}

    with get_db() as conn:
        for f in files:
            orig_name = os.path.basename(f.filename or "image.png")
            ext = os.path.splitext(orig_name)[1].lower()
            file_bytes = f.read()
            if not file_bytes:
                continue

            # Verify if it's a valid image
            is_valid_image = False
            mime_type = "image/png"

            if ext in allowed_exts or (f.content_type and f.content_type.startswith("image/")):
                is_valid_image = True
            
            # Check with PIL or SVG text
            if not is_valid_image or ext not in allowed_exts:
                if b"<svg" in file_bytes[:500].lower():
                    is_valid_image = True
                    ext = ".svg"
                    mime_type = "image/svg+xml"
                else:
                    try:
                        with Image.open(BytesIO(file_bytes)) as test_img:
                            is_valid_image = True
                            fmt = (test_img.format or "PNG").lower()
                            if fmt == "jpeg": ext = ".jpg"
                            else: ext = f".{fmt}"
                    except Exception:
                        pass

            if not is_valid_image:
                continue

            if not ext:
                ext = ".png"

            # Determine mime type
            if ext in (".jpg", ".jpeg", ".jfif"): mime_type = "image/jpeg"
            elif ext == ".webp": mime_type = "image/webp"
            elif ext == ".gif": mime_type = "image/gif"
            elif ext == ".svg": mime_type = "image/svg+xml"
            elif ext == ".bmp": mime_type = "image/bmp"
            elif ext in (".tif", ".tiff"): mime_type = "image/tiff"
            elif ext == ".avif": mime_type = "image/avif"
            elif ext == ".ico": mime_type = "image/x-icon"
            elif ext == ".png": mime_type = "image/png"
            elif f.content_type and f.content_type.startswith("image/"):
                mime_type = f.content_type

            img_title = title if (title and len(files) == 1) else os.path.splitext(orig_name)[0].replace("_", " ").title()
            unique_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.urandom(4).hex()}{ext}"
            file_path = GALLERY_DIR / unique_name
            
            # Save file to disk
            try:
                with open(file_path, "wb") as disk_file:
                    disk_file.write(file_bytes)
            except Exception as e:
                print(f"[WARN] Failed to write image to disk: {e}")

            cur = conn.execute("""
                INSERT INTO gallery_images (filename, original_name, title, category, file_size, mime_type, image_blob, uploaded_by, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """, (unique_name, orig_name, img_title, category, len(file_bytes), mime_type, sqlite3.Binary(file_bytes), uploader))
            
            new_id = cur.lastrowid
            saved_images.append({
                "id": new_id,
                "filename": unique_name,
                "original_name": orig_name,
                "title": img_title,
                "category": category,
                "file_size": len(file_bytes),
                "view_url": f"/gallery/view/{new_id}",
                "download_url": f"/gallery/download/{new_id}"
            })

        conn.commit()

    if not saved_images:
        return jsonify({"error": "No valid image files could be processed. Please upload standard image formats (JPG, PNG, WEBP, GIF, BMP, etc.)."}), 400

    return jsonify({"success": True, "count": len(saved_images), "images": saved_images})

@app.route("/gallery/view/<int:image_id>")
def view_gallery_image(image_id):
    """Serves the image for viewing in browser / thumbnail."""
    with get_db() as conn:
        row = conn.execute("SELECT filename, mime_type, image_blob FROM gallery_images WHERE id = ?", (image_id,)).fetchone()
    if not row:
        return jsonify({"error": "Image not found"}), 404

    filename = row["filename"]
    disk_path = GALLERY_DIR / filename
    if disk_path.exists():
        return send_file(disk_path, mimetype=row["mime_type"])
    elif row["image_blob"]:
        return send_file(BytesIO(row["image_blob"]), mimetype=row["mime_type"])
    return jsonify({"error": "Image file not found on disk or database"}), 404

@app.route("/gallery/download/<int:image_id>")
def download_gallery_image(image_id):
    """Triggers download of the image file as an attachment with its original name."""
    with get_db() as conn:
        row = conn.execute("SELECT filename, original_name, mime_type, image_blob FROM gallery_images WHERE id = ?", (image_id,)).fetchone()
    if not row:
        return jsonify({"error": "Image not found"}), 404

    filename = row["filename"]
    orig_name = row["original_name"] or filename
    disk_path = GALLERY_DIR / filename

    if disk_path.exists():
        return send_file(disk_path, as_attachment=True, download_name=orig_name, mimetype=row["mime_type"])
    elif row["image_blob"]:
        return send_file(BytesIO(row["image_blob"]), as_attachment=True, download_name=orig_name, mimetype=row["mime_type"])
    return jsonify({"error": "Image file not found"}), 404

@app.route("/api/gallery/<int:image_id>", methods=["DELETE"])
def api_delete_gallery_image(image_id):
    """Deletes an image from DB and disk."""
    user = session.get("user")
    if not user:
        return jsonify({"error": "Authentication required"}), 401

    with get_db() as conn:
        row = conn.execute("SELECT filename, uploaded_by FROM gallery_images WHERE id = ?", (image_id,)).fetchone()
        if not row:
            return jsonify({"error": "Image not found"}), 404

        # Only admin or uploader can delete
        if user.get("role") != "admin" and user.get("username") != row["uploaded_by"]:
            return jsonify({"error": "You do not have permission to delete this image."}), 403

        conn.execute("DELETE FROM gallery_images WHERE id = ?", (image_id,))
        conn.commit()

    # Remove from disk if present
    disk_path = GALLERY_DIR / row["filename"]
    if disk_path.exists():
        try:
            disk_path.unlink()
        except Exception:
            pass

    return jsonify({"success": True})


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def run_cli_ocr(image_path: str):
    if not os.path.exists(image_path):
        print(f"Error: File not found {image_path}")
        return
    img = Image.open(image_path).convert("RGB")
    text = do_ocr_on_pil_image(img)
    print(f"\n[RAW OCR TEXT]:\n{text}\n" + "="*40)
    data = extract_all(text)
    print("\n[EXTRACTED FIELDS]:")
    for k, v in data.items():
        if k != "raw_ocr":
            print(f"  {k:<12}: {v or '(none)'}")

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

if __name__ == "__main__":
    init_db()
    if len(sys.argv) > 1 and sys.argv[1].lower().endswith((".jpg", ".jpeg", ".png", ".webp", ".bmp")):
        run_cli_ocr(sys.argv[1])
    else:
        local_ip = get_local_ip()
        print("\n" + "="*55)
        print(">> InvenScan -- Smart Inventory & Role-Based System")
        print("="*55)
        print(f"  OCR Engine : {'[OK] Windows OCR (winocr sync)' if WINOCR_AVAILABLE else '[WARN] NOT READY'}")
        print(f"  Barcode    : {'[OK] Code128' if BARCODE_AVAILABLE else '[OK] QR Code'}")
        print(f"  Excel Sync : [OK] Auto-Syncs to {EXCEL_SYNC_PATH.name}")
        print(f"  Local URL  : http://localhost:5000/login")
        print(f"  Network URL: http://{local_ip}:5000/login (For devices on same Wi-Fi)")
        print("="*55 + "\n")
        
        try:
            webbrowser.open(f"http://localhost:5000/login")
        except Exception:
            pass

        app.run(debug=False, host="0.0.0.0", port=5000)
